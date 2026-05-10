#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, DefaultDict
from collections import defaultdict

import rispy  # rispy.load parses .ris -> list[dict] [1](https://pypi.org/project/rispy/)
from openpyxl import Workbook, load_workbook

import tkinter as tk
from tkinter import filedialog, messagebox  # dialogs like askopenfilenames/askdirectory [2](https://docs.python.org/3/library/dialog.html)


# ----------------------------
# Defaults
# ----------------------------

DEFAULT_SHEET = "Archive"
PROCESSED_SHEET = "__ProcessedRIS"

DEFAULT_COLUMNS = [
    "Author",
    "Title",
    "Year of Publication",
    "Publisher",
    "Location",
    "URL (if available)",
    "PDF name",
    "Text file name",
    "Duplicate?",
    "Duplicate Key",
]


# ----------------------------
# Duplicate fingerprinting
# ----------------------------

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def make_fingerprint(row: Dict[str, str]) -> str:
    """
    Fingerprint used to flag duplicates (but never remove them).
    Prefer URL if present; otherwise Title+Author+Year.
    """
    url = _norm(row.get("URL (if available)", ""))
    title = _norm(row.get("Title", ""))
    author = _norm(row.get("Author", ""))
    year = _norm(row.get("Year of Publication", ""))
    if url:
        return f"url:{url}"
    return f"t:{title}|a:{author}|y:{year}"


# ----------------------------
# RIS parsing
# ----------------------------

def _first_present(entry: Dict[str, Any], keys: Sequence[str]) -> Optional[Any]:
    for k in keys:
        if k in entry and entry[k] not in (None, "", []):
            return entry[k]
    return None


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(x) for x in value if x is not None)
    return str(value)


def _authors(entry: Dict[str, Any]) -> str:
    v = _first_present(entry, ["first_authors", "authors"])
    return _as_text(v)


def _title(entry: Dict[str, Any]) -> str:
    v = _first_present(entry, ["primary_title", "title", "secondary_title", "tertiary_title"])
    return _as_text(v)


def _publisher(entry: Dict[str, Any]) -> str:
    v = _first_present(entry, ["publisher"])
    return _as_text(v)


def _location(entry: Dict[str, Any]) -> str:
    v = _first_present(entry, ["place_published", "city", "publication_place"])
    return _as_text(v)


def _url(entry: Dict[str, Any]) -> str:
    v = _first_present(entry, ["url", "urls"])
    if isinstance(v, list):
        return str(v[0]) if v else ""
    return _as_text(v)


def _year_raw(entry: Dict[str, Any]) -> str:
    """
    Preserve whatever RIS provides as year/date (ranges, conjectures, seasons, etc.).
    """
    candidate_keys = [
        "year",
        "publication_year",
        "date",
        "publication_date",
        "primary_date",
        "secondary_date",
        "y1",
        "py",
    ]
    v = _first_present(entry, candidate_keys)

    if v is None:
        for k in entry.keys():
            kl = k.lower()
            if ("date" in kl) or ("year" in kl):
                vv = entry.get(k)
                if vv not in (None, "", []):
                    v = vv
                    break

    return _as_text(v)


def parse_ris_file(ris_path: Path) -> List[Dict[str, str]]:
    with ris_path.open("r", encoding="utf-8", errors="replace") as f:
        entries = rispy.load(f)  # [1](https://pypi.org/project/rispy/)

    out: List[Dict[str, str]] = []
    for e in entries:
        out.append({
            "Author": _authors(e),
            "Title": _title(e),
            "Year of Publication": _year_raw(e),
            "Publisher": _publisher(e),
            "Location": _location(e),
            "URL (if available)": _url(e),
        })
    return out


# ----------------------------
# PDF association by timestamp
# ----------------------------

@dataclass(frozen=True)
class PdfFile:
    path: Path
    mtime: float


def list_pdfs(folder: Path) -> List[PdfFile]:
    pdfs: List[PdfFile] = []
    for p in folder.glob("*.pdf"):
        try:
            pdfs.append(PdfFile(p, p.stat().st_mtime))
        except OSError:
            pass
    pdfs.sort(key=lambda x: x.mtime)
    return pdfs


def match_pdfs_to_entries(ris_mtime: float, entry_count: int, pdfs: Sequence[PdfFile]) -> List[str]:
    """
    Choose PDFs modified soonest AFTER RIS mtime.
    - candidates: pdf.mtime > ris_mtime
    - if one: use for all
    - if many: map chronologically; if fewer PDFs than entries, repeat last
    """
    candidates = [p for p in pdfs if p.mtime > ris_mtime]
    if not candidates:
        return [""] * entry_count
    if len(candidates) == 1:
        return [candidates[0].path.name] * entry_count

    names: List[str] = []
    for i in range(entry_count):
        idx = min(i, len(candidates) - 1)
        names.append(candidates[idx].path.name)
    return names


# ----------------------------
# Text file association (FIXED: RIS first, TXT after; one-to-one)
# ----------------------------

@dataclass(frozen=True)
class TextFile:
    path: Path
    mtime: float


def list_text_files(folder: Path) -> List[TextFile]:
    """
    Treat 'plain text files' as .txt/.TXT.
    """
    out: List[TextFile] = []
    for patt in ("*.txt", "*.TXT"):
        for p in folder.glob(patt):
            try:
                out.append(TextFile(p, p.stat().st_mtime))
            except OSError:
                pass
    out.sort(key=lambda x: x.mtime)
    return out


def compute_next_ris_mtime_map(ris_files: Sequence[Path]) -> Dict[Path, Optional[float]]:
    """
    For RIS files within the same folder, compute each file's 'next RIS mtime'
    by sorting by mtime. Returns map: ris_path -> next_ris_mtime (or None).
    """
    by_folder: DefaultDict[Path, List[Tuple[Path, float]]] = defaultdict(list)
    for p in ris_files:
        try:
            by_folder[p.parent].append((p, p.stat().st_mtime))
        except OSError:
            pass

    next_map: Dict[Path, Optional[float]] = {}
    for folder, items in by_folder.items():
        items.sort(key=lambda t: t[1])
        for i, (p, mt) in enumerate(items):
            nxt = items[i + 1][1] if i + 1 < len(items) else None
            next_map[p] = nxt
    return next_map


def build_text_map_after_ris(
    ris_files: Sequence[Path],
    window_seconds: int = 300,
) -> Dict[Path, str]:
    """
    Assign at most ONE .txt to each .ris (one-to-one, no reuse), where:

      - RIS is downloaded first, then TXT
      - Eligible TXT must satisfy:
          * txt.mtime > ris.mtime
          * AND (
                (next RIS exists AND txt.mtime < next_ris.mtime)  OR
                (no next RIS AND txt.mtime <= ris.mtime + window_seconds)
              )

    If multiple TXT are eligible for a given RIS, choose the one with the smallest (txt.mtime - ris.mtime).
    Ensures each TXT is used at most once across the folder.
    """
    next_map = compute_next_ris_mtime_map(ris_files)

    # Group RIS by folder, with their mtimes
    by_folder: DefaultDict[Path, List[Tuple[Path, float]]] = defaultdict(list)
    for r in ris_files:
        try:
            by_folder[r.parent].append((r, r.stat().st_mtime))
        except OSError:
            pass

    result: Dict[Path, str] = {}

    for folder, ris_items in by_folder.items():
        ris_items.sort(key=lambda t: t[1])  # chronological
        texts = list_text_files(folder)
        if not texts:
            continue

        used_txt: set[str] = set()

        # For each RIS in chronological order, pick closest TXT after it within its eligibility window
        for ris_path, ris_mt in ris_items:
            nxt = next_map.get(ris_path)
            best_name = ""
            best_delta: Optional[float] = None

            for tf in texts:
                if tf.path.name in used_txt:
                    continue
                if tf.mtime <= ris_mt:
                    continue

                eligible = False
                if nxt is not None:
                    # anything strictly between RIS and next RIS
                    if tf.mtime < nxt:
                        eligible = True
                else:
                    # last RIS: within 5 minutes after
                    if tf.mtime <= (ris_mt + window_seconds):
                        eligible = True

                if not eligible:
                    continue

                delta = tf.mtime - ris_mt
                if best_delta is None or delta < best_delta:
                    best_delta = delta
                    best_name = tf.path.name

            if best_name:
                result[ris_path] = best_name
                used_txt.add(best_name)

    return result


# ----------------------------
# Excel helpers (append-only)
# ----------------------------

def _ensure_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _read_headers(ws) -> List[str]:
    headers: List[str] = []
    for cell in ws[1]:
        if cell.value is None or str(cell.value).strip() == "":
            break
        headers.append(str(cell.value))
    return headers


def _write_headers(ws, headers: Sequence[str]) -> None:
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)


def _iter_sheet_rows_as_dicts(ws, headers: Sequence[str]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    if not headers:
        return out
    col_index = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        d: Dict[str, str] = {}
        empty = True
        for h in headers:
            v = ws.cell(row=r, column=col_index[h]).value
            s = "" if v is None else str(v)
            if s.strip():
                empty = False
            d[h] = s
        if not empty:
            out.append(d)
    return out


def _load_or_create_workbook(xlsx_path: Path):
    if xlsx_path.exists():
        return load_workbook(xlsx_path)
    return Workbook()


def append_rows_to_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    rows: List[Dict[str, str]],
    columns: Sequence[str],
    do_flag_duplicates: bool = True,
) -> None:
    wb = _load_or_create_workbook(xlsx_path)
    ws = _ensure_sheet(wb, sheet_name)
    _ensure_sheet(wb, PROCESSED_SHEET)

    existing_headers = _read_headers(ws)
    if not existing_headers:
        _write_headers(ws, columns)
        existing_headers = list(columns)
    else:
        missing = [c for c in columns if c not in existing_headers]
        if missing:
            existing_headers = existing_headers + missing
            _write_headers(ws, existing_headers)

    existing_fps = set()
    if do_flag_duplicates:
        for er in _iter_sheet_rows_as_dicts(ws, existing_headers):
            existing_fps.add(make_fingerprint(er))

        for dc in ("Duplicate?", "Duplicate Key"):
            if dc not in existing_headers:
                existing_headers.append(dc)
                _write_headers(ws, existing_headers)

    col_index = {h: i + 1 for i, h in enumerate(existing_headers)}
    start_row = ws.max_row + 1

    for i, row in enumerate(rows):
        if do_flag_duplicates:
            fp = make_fingerprint(row)
            row["Duplicate Key"] = fp
            row["Duplicate?"] = "YES" if fp in existing_fps else "NO"
            existing_fps.add(fp)

        r = start_row + i
        for c in columns:
            ws.cell(row=r, column=col_index[c], value=row.get(c, ""))

        if do_flag_duplicates:
            for dc in ("Duplicate?", "Duplicate Key"):
                if dc in col_index:
                    ws.cell(row=r, column=col_index[dc], value=row.get(dc, ""))

    wb.save(xlsx_path)


# ----------------------------
# Skip-already-processed tracking
# ----------------------------

def _processed_key(ris_path: Path) -> Tuple[str, int, int]:
    st = ris_path.stat()
    return (ris_path.name, int(st.st_mtime), int(st.st_size))


def load_processed_keys(xlsx_path: Path) -> set:
    if not xlsx_path.exists():
        return set()
    wb = load_workbook(xlsx_path)
    if PROCESSED_SHEET not in wb.sheetnames:
        return set()
    ws = wb[PROCESSED_SHEET]
    headers = _read_headers(ws)
    if not headers:
        return set()
    keys = set()
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        mtime = ws.cell(row=r, column=2).value
        size = ws.cell(row=r, column=3).value
        if name and mtime and size:
            try:
                keys.add((str(name), int(float(mtime)), int(float(size))))
            except Exception:
                pass
    return keys


def append_processed_keys(xlsx_path: Path, keys_to_add: List[Tuple[str, int, int]]) -> None:
    wb = _load_or_create_workbook(xlsx_path)
    ws = _ensure_sheet(wb, PROCESSED_SHEET)
    headers = _read_headers(ws)
    if not headers:
        _write_headers(ws, ["RIS name", "mtime", "size", "processed_at_epoch"])
    start_row = ws.max_row + 1
    now = int(time.time())
    for i, (name, mtime, size) in enumerate(keys_to_add):
        r = start_row + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=mtime)
        ws.cell(row=r, column=3, value=size)
        ws.cell(row=r, column=4, value=now)
    wb.save(xlsx_path)


# ----------------------------
# Processing pipeline
# ----------------------------

def discover_ris_in_folder(folder: Path) -> List[Path]:
    return sorted(folder.glob("*.ris")) + sorted(folder.glob("*.RIS"))


def resolve_output_xlsx(
    input_folder: Optional[Path],
    outputfolder: Optional[Path],
    xlsx_arg: Optional[str],
    default_name: str,
) -> Path:
    """
    Rules:
    - If xlsx_arg is an explicit path (has a parent dir), use it.
    - Else if outputfolder provided, create there.
    - Else if input_folder provided, create there.
    - Else create in current working directory.
    """
    if xlsx_arg:
        p = Path(xlsx_arg)
        if str(p.parent) not in (".", ""):
            return p
        name = p.name
    else:
        name = default_name

    if outputfolder:
        return outputfolder / name
    if input_folder:
        return input_folder / name
    return Path.cwd() / name


def process_ris_files(
    ris_files: Sequence[Path],
    xlsx_path: Path,
    columns: Sequence[str],
    sheet_name: str,
    auto_pdf: bool,
    flag_duplicates: bool,
    skip_processed: bool,
    gui_parent: Optional[tk.Tk] = None,
) -> Tuple[int, int, int]:
    cols = list(columns)
    if flag_duplicates:
        for dc in ("Duplicate?", "Duplicate Key"):
            if dc not in cols:
                cols.append(dc)

    processed_keys = load_processed_keys(xlsx_path) if skip_processed else set()

    # Build one-to-one text mapping for this batch
    text_map = build_text_map_after_ris(ris_files, window_seconds=300) if "Text file name" in cols else {}

    ris_processed = 0
    rows_appended = 0
    ris_skipped = 0
    newly_processed: List[Tuple[str, int, int]] = []

    for ris_path in ris_files:
        if not ris_path.exists():
            continue

        key = _processed_key(ris_path)
        if skip_processed and key in processed_keys:
            ris_skipped += 1
            continue

        entries = parse_ris_file(ris_path)
        folder = ris_path.parent
        ris_mtime = ris_path.stat().st_mtime

        # Text assignment (same for all entries in this RIS)
        text_name = text_map.get(ris_path, "") if "Text file name" in cols else ""

        # PDF matching
        pdf_names = [""] * len(entries)
        if "PDF name" in cols and entries:
            if auto_pdf:
                pdfs = list_pdfs(folder)
                pdf_names = match_pdfs_to_entries(ris_mtime, len(entries), pdfs)
            else:
                if gui_parent is not None:
                    chosen = filedialog.askopenfilenames(
                        parent=gui_parent,
                        title=f"Select PDF(s) for {ris_path.name}",
                        initialdir=str(folder),
                        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                    )  # [2](https://docs.python.org/3/library/dialog.html)
                    chosen_list = [Path(p).name for p in chosen]
                    if len(chosen_list) == 1:
                        pdf_names = [chosen_list[0]] * len(entries)
                    elif len(chosen_list) >= len(entries):
                        pdf_names = chosen_list[:len(entries)]
                    elif len(chosen_list) > 1:
                        pdf_names = [chosen_list[min(i, len(chosen_list) - 1)] for i in range(len(entries))]

        rows: List[Dict[str, str]] = []
        for i, e in enumerate(entries):
            row = dict(e)
            if "PDF name" in cols:
                row["PDF name"] = pdf_names[i] if i < len(pdf_names) else ""
            if "Text file name" in cols:
                row["Text file name"] = text_name
            for c in cols:
                row.setdefault(c, "")
            rows.append(row)

        if rows:
            append_rows_to_xlsx(
                xlsx_path=xlsx_path,
                sheet_name=sheet_name,
                rows=rows,
                columns=cols,
                do_flag_duplicates=flag_duplicates,
            )
            rows_appended += len(rows)

        newly_processed.append(key)
        ris_processed += 1

    if newly_processed:
        append_processed_keys(xlsx_path, newly_processed)

    return ris_processed, rows_appended, ris_skipped


# ----------------------------
# GUI
# ----------------------------

class RisArchiverGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RIS Archiver (Append to Excel)")
        self.geometry("1100x650")
        self.resizable(True, True)

        self.grid_columnconfigure(1, weight=1)

        self.selected_ris_files: List[Path] = []

        self.input_folder_var = tk.StringVar(value="")
        self.output_folder_var = tk.StringVar(value="")
        self.xlsx_name_var = tk.StringVar(value="archive.xlsx")
        self.xlsx_fullpath_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value=DEFAULT_SHEET)

        self.auto_pdf_var = tk.BooleanVar(value=True)
        self.flag_dupes_var = tk.BooleanVar(value=True)
        self.skip_processed_var = tk.BooleanVar(value=True)
        self.columns_var = tk.StringVar(value=",".join(DEFAULT_COLUMNS))

        self._build()

    def _build(self):
        pad = {"padx": 10, "pady": 6}

        top_frame = tk.Frame(self)
        top_frame.grid(row=0, column=0, columnspan=3, sticky="we", padx=10, pady=10)
        top_frame.grid_columnconfigure(2, weight=1)

        tk.Button(top_frame, text="Select .RIS files…", command=self.select_ris_files).grid(row=0, column=0, padx=6)
        tk.Button(top_frame, text="Select Input Folder…", command=self.pick_input_folder).grid(row=0, column=1, padx=6)

        self.selected_ris_label = tk.Label(top_frame, text="No RIS files selected.", fg="#444", anchor="w")
        self.selected_ris_label.grid(row=0, column=2, padx=10, sticky="we")

        tk.Label(self, text="Input folder (RIS/PDF/TXT):").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.input_folder_var).grid(row=1, column=1, sticky="we", **pad)

        tk.Label(self, text="Output folder (optional):").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.output_folder_var).grid(row=2, column=1, sticky="we", **pad)
        tk.Button(self, text="Browse…", command=self.pick_output_folder).grid(row=2, column=2, **pad)

        tk.Label(self, text="Output .xlsx filename:").grid(row=3, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.xlsx_name_var, width=30).grid(row=3, column=1, sticky="w", **pad)
        tk.Button(self, text="Choose existing .xlsx…", command=self.pick_existing_xlsx).grid(row=3, column=2, **pad)

        tk.Label(self, text="(Choosing an existing .xlsx overrides output folder + filename.)",
                 fg="#444").grid(row=4, column=1, sticky="w", padx=10, pady=2)

        tk.Label(self, text="Sheet name:").grid(row=5, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.sheet_var, width=24).grid(row=5, column=1, sticky="w", **pad)

        tk.Checkbutton(self, text="Auto-associate PDFs (timestamp: soonest AFTER RIS time)",
                       variable=self.auto_pdf_var).grid(row=6, column=0, columnspan=3, sticky="w", **pad)

        tk.Checkbutton(self, text="Flag duplicates (append anyway; mark Duplicate? and Duplicate Key)",
                       variable=self.flag_dupes_var).grid(row=7, column=0, columnspan=3, sticky="w", **pad)

        tk.Checkbutton(self, text="Skip already-processed RIS files (tracked in workbook metadata sheet)",
                       variable=self.skip_processed_var).grid(row=8, column=0, columnspan=3, sticky="w", **pad)

        tk.Label(self, text="Columns (comma-separated):").grid(row=9, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.columns_var).grid(row=9, column=1, sticky="we", **pad)
        tk.Button(self, text="Reset columns", command=self.reset_columns).grid(row=9, column=2, **pad)

        btn_frame = tk.Frame(self)
        btn_frame.grid(row=10, column=1, sticky="w", **pad)
        tk.Button(btn_frame, text="Run on Folder", command=self.run_folder).grid(row=0, column=0, padx=6)

        self.status = tk.Label(self, text="Ready.", anchor="w")
        self.status.grid(row=11, column=0, columnspan=3, sticky="we", padx=10, pady=14)

        tip = (
            "Tip: Close the .xlsx in Excel/Numbers before running.\n"
            "Year field is preserved exactly as provided by the RIS export.\n"
            "Duplicate flagging never removes rows; it only marks them.\n"
            "TXT matching (fixed): RIS is downloaded first; we match ONE .txt created AFTER the RIS.\n"
            "If there is a next RIS in the folder, the TXT must fall before it. If this RIS is last, TXT must be within 5 minutes.\n"
        )
        tk.Label(self, text=tip, justify="left", fg="#444", wraplength=1050).grid(
            row=12, column=0, columnspan=3, sticky="w", padx=10
        )

    def select_ris_files(self):
        ris_paths = filedialog.askopenfilenames(
            title="Select .ris file(s)",
            filetypes=[("RIS files", "*.ris *.RIS"), ("All files", "*.*")],
        )  # [2](https://docs.python.org/3/library/dialog.html)
        if not ris_paths:
            return

        self.selected_ris_files = [Path(p) for p in ris_paths]
        folder = self.selected_ris_files[0].parent
        self.input_folder_var.set(str(folder))
        self.selected_ris_label.config(text=f"Selected RIS files: {len(self.selected_ris_files)}")
        self.run_selected_ris()

    def pick_input_folder(self):
        d = filedialog.askdirectory(title="Select input folder (RIS/PDF/TXT)")  # [2](https://docs.python.org/3/library/dialog.html)
        if d:
            self.input_folder_var.set(d)
            self.selected_ris_files = []
            self.selected_ris_label.config(text="No RIS files selected.")

    def pick_output_folder(self):
        d = filedialog.askdirectory(title="Select output folder (optional)")  # [2](https://docs.python.org/3/library/dialog.html)
        if d:
            self.output_folder_var.set(d)

    def pick_existing_xlsx(self):
        f = filedialog.askopenfilename(
            title="Choose existing Excel workbook",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )  # [2](https://docs.python.org/3/library/dialog.html)
        if f:
            self.xlsx_fullpath_var.set(f)

    def reset_columns(self):
        self.columns_var.set(",".join(DEFAULT_COLUMNS))

    def _get_columns(self) -> List[str]:
        cols = [c.strip() for c in self.columns_var.get().split(",") if c.strip()]
        return cols if cols else list(DEFAULT_COLUMNS)

    def _resolve_xlsx_path(self, input_folder: Path) -> Path:
        chosen = self.xlsx_fullpath_var.get().strip()
        if chosen:
            return Path(chosen)

        out_folder = self.output_folder_var.get().strip()
        out_folder_path = Path(out_folder) if out_folder else input_folder

        name = self.xlsx_name_var.get().strip() or "archive.xlsx"
        if not name.lower().endswith(".xlsx"):
            name += ".xlsx"
        return out_folder_path / name

    def run_selected_ris(self):
        if not self.selected_ris_files:
            return

        input_folder = self.selected_ris_files[0].parent
        xlsx_path = self._resolve_xlsx_path(input_folder)
        sheet = (self.sheet_var.get().strip() or DEFAULT_SHEET)
        cols = self._get_columns()

        self.status.config(text=f"Processing {len(self.selected_ris_files)} selected RIS file(s)…")
        self.update_idletasks()

        try:
            rp, ra, rs = process_ris_files(
                ris_files=self.selected_ris_files,
                xlsx_path=xlsx_path,
                columns=cols,
                sheet_name=sheet,
                auto_pdf=self.auto_pdf_var.get(),
                flag_duplicates=self.flag_dupes_var.get(),
                skip_processed=self.skip_processed_var.get(),
                gui_parent=self,
            )
        except PermissionError:
            messagebox.showerror(
                "Excel file in use",
                "Could not save the .xlsx. Please close it in Excel/Numbers and try again."
            )
            return

        self.status.config(
            text=f"Done. RIS processed: {rp}. RIS skipped: {rs}. Rows appended: {ra}. Output: {xlsx_path}"
        )

    def run_folder(self):
        inp = self.input_folder_var.get().strip()
        if not inp:
            messagebox.showerror("Missing input", "Please select an input folder.")
            return

        input_folder = Path(inp)
        ris_files = discover_ris_in_folder(input_folder)
        if not ris_files:
            messagebox.showinfo("No RIS files", "No .ris files found in that folder.")
            return

        xlsx_path = self._resolve_xlsx_path(input_folder)
        sheet = (self.sheet_var.get().strip() or DEFAULT_SHEET)
        cols = self._get_columns()

        self.status.config(text=f"Processing {len(ris_files)} RIS file(s)…")
        self.update_idletasks()

        try:
            rp, ra, rs = process_ris_files(
                ris_files=ris_files,
                xlsx_path=xlsx_path,
                columns=cols,
                sheet_name=sheet,
                auto_pdf=self.auto_pdf_var.get(),
                flag_duplicates=self.flag_dupes_var.get(),
                skip_processed=self.skip_processed_var.get(),
                gui_parent=self,
            )
        except PermissionError:
            messagebox.showerror(
                "Excel file in use",
                "Could not save the .xlsx. Please close it in Excel/Numbers and try again."
            )
            return

        self.status.config(
            text=f"Done. RIS processed: {rp}. RIS skipped: {rs}. Rows appended: {ra}. Output: {xlsx_path}"
        )


# ----------------------------
# CLI
# ----------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="ris-archiver",
        description="Archive RIS citations into an Excel workbook (append-only), with PDF/TXT association, duplicate flagging, and skip-processed.",
    )
    src = p.add_mutually_exclusive_group()
    src.add_argument("--folder", type=str, help="Folder containing .ris and .pdf/.txt files.")
    src.add_argument("--ris", nargs="+", help="One or more .ris files to process.")

    p.add_argument("--xlsx", type=str, default=None,
                   help="Output .xlsx path OR filename (default: archive.xlsx in outputfolder/input folder).")
    p.add_argument("--outputfolder", type=str, default=None,
                   help="Output folder for .xlsx when --xlsx is a filename or omitted.")
    p.add_argument("--sheet", type=str, default=DEFAULT_SHEET, help=f"Sheet name (default: {DEFAULT_SHEET}).")

    p.add_argument("--auto-pdf", dest="auto_pdf", action="store_true",
                   help="Auto-associate PDFs by timestamp (soonest AFTER RIS time).")
    p.add_argument("--no-auto-pdf", dest="auto_pdf", action="store_false",
                   help="Disable auto PDF association.")
    p.set_defaults(auto_pdf=True)

    p.add_argument("--flag-duplicates", dest="flag_duplicates", action="store_true",
                   help="Flag duplicates (append anyway; mark Duplicate? and Duplicate Key).")
    p.add_argument("--no-flag-duplicates", dest="flag_duplicates", action="store_false",
                   help="Disable duplicate flagging.")
    p.set_defaults(flag_duplicates=True)

    p.add_argument("--skip-processed", dest="skip_processed", action="store_true",
                   help="Skip RIS files already logged as processed in workbook metadata sheet.")
    p.add_argument("--no-skip-processed", dest="skip_processed", action="store_false",
                   help="Do not skip processed RIS files.")
    p.set_defaults(skip_processed=True)

    p.add_argument("--columns", type=str, default=",".join(DEFAULT_COLUMNS),
                   help="Comma-separated column list.")

    return p


def main(argv: Sequence[str]) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv))

    if not any([args.folder, args.ris, args.xlsx, args.outputfolder]) and len(argv) == 0:
        gui = RisArchiverGUI()
        gui.mainloop()
        return 0

    outfolder = Path(args.outputfolder) if args.outputfolder else None

    input_folder: Optional[Path] = None
    if args.folder:
        input_folder = Path(args.folder)
        ris_files = discover_ris_in_folder(input_folder)
    elif args.ris:
        ris_files = [Path(p) for p in args.ris]
        input_folder = ris_files[0].parent if ris_files else None
    else:
        print("ERROR: Provide --folder or --ris in CLI mode (or run with no args for GUI).")
        return 2

    if not ris_files:
        print("No RIS files found.")
        return 0

    xlsx_path = resolve_output_xlsx(
        input_folder=input_folder,
        outputfolder=outfolder,
        xlsx_arg=args.xlsx,
        default_name="archive.xlsx",
    )

    columns = [c.strip() for c in (args.columns or "").split(",") if c.strip()]
    if not columns:
        columns = list(DEFAULT_COLUMNS)

    rp, ra, rs = process_ris_files(
        ris_files=ris_files,
        xlsx_path=xlsx_path,
        columns=columns,
        sheet_name=args.sheet or DEFAULT_SHEET,
        auto_pdf=args.auto_pdf,
        flag_duplicates=args.flag_duplicates,
        skip_processed=args.skip_processed,
        gui_parent=None,
    )

    print(f"Done. RIS processed: {rp}. RIS skipped: {rs}. Rows appended: {ra}. Output: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))